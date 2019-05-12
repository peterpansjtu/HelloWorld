import os
import time

import openpyxl


class PLCResult(object):
    def __init__(self, path):
        self.path = path
        self.time_column = 1
        self.bar_code_column = 2
        self.pumping_time_column = 3
        self.pressure_column = 4
        self.test_result_column = 5
        self.next_row = 1
        if os.path.isfile(path):
            self.wb = openpyxl.load_workbook(path)
            self.ws = self.wb.get_sheet_by_name('PLC结果')
            self.next_row = self.ws.max_row + 1
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "PLC结果"
            self.ws.cell(1, self.time_column, "时间")
            self.ws.cell(1, self.bar_code_column, "条形码")
            self.ws.cell(1, self.pumping_time_column, "抽气时间")
            self.ws.cell(1, self.pressure_column, "负压(Kpa)")
            self.ws.cell(1, self.test_result_column, "测试结果")
            self.next_row += 1

    def add_result(self, pumping_time, pressure, test_result, bar_code=''):
        self.ws.cell(self.next_row, self.time_column, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        self.ws.cell(self.next_row, self.bar_code_column, bar_code)
        self.ws.cell(self.next_row, self.pumping_time_column, pumping_time)
        self.ws.cell(self.next_row, self.pressure_column, pressure)
        self.ws.cell(self.next_row, self.test_result_column, test_result)
        self.next_row += 1
        self.wb.save(self.path)


if __name__ == '__main__':
    result = PLCResult('1.xlsx')
    # result.add_result(0.000001, 3, 'ok')
    # result.add_result(0.000002, 4, 'ok')
